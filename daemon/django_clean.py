# -*- coding: utf-8 -*-
from __future__ import unicode_literals

#import pyinotify
#import signal
#import threading
#import shutil
#import posixpath
#import time
import sys
import os
import traceback
import openpyxl

APPS_PATH = "/var/www/django/prod"
if __name__ == '__main__':
	sys.path.insert(0, APPS_PATH)
	os.chdir(APPS_PATH)
	os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
	import django
	django.setup()


MARGE_PERDU = 180					# 180 jours après réception
MARGE_DESTRUCTION = 30				# 30 jours après fin de validité
MARGE_SUPPRESSION = -7				# 7 jours avant fin de validité
MARGE_NOUVELLE_COLLECTION = 30 		# Mouvement collection moins de 30 jours
	
from fonction import *
from stock_labo.models import Contenant, Nomenclature, NomenclatureLot
from lib_linautom.python import mail
from django.contrib.auth.models import Permission, User
from django.conf import settings
from lib_linautom.python import mysql


if version.status_developement["data"] == "prod":
	fichier_log = "/var/log/django_clean_prod.log"
	XLSX_PATH = "/var/backups/django_clean"
else:
	fichier_log = "/var/log/django_clean_dev.log"
	XLSX_PATH = "/var/backups/django_clean"


#Affichage des log sur la console
console = False
console = True

class DossierInconnu(Exception):
	"""
	Le dossier est inconnu
	"""
	def __init__(self, *args, **kwargs):
		self.args = args
		self.kwargs = kwargs
	def __str__(self):
		print_ascii(__doc__)

def log_print(*text):
	"""
	Fonction de login automatique
	"""
	f_log = codecs.open(fichier_log, "a", encoding="utf-8")
	d = datetime.datetime.now().strftime("%a, %Y-%m-%d %X ")
	for t in text:
		if console == True:
			print_ascii(t)
		l = d + unicode(t) + "\n"
		f_log.writelines(l)
	f_log.close()


def construction_page_xlsx(f, liste, titre):
	"""
	Construction fichier .xlsx
	f : feuille d'un fichier xlsx
	liste : liste de donnée
	titre : titre de la page
	"""
	f.merge_cells("A1:H1")
	column_dimension = []
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[0].auto_size = True
	column_dimension[0].width = 1/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(0+1): column_dimension[0]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[1].auto_size = True
	column_dimension[1].width = 4/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(1+1): column_dimension[1]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[2].auto_size = True
	column_dimension[2].width = 5/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(2+1): column_dimension[2]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[3].auto_size = True
	column_dimension[3].width = 15/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(3+1): column_dimension[3]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[4].auto_size = True
	column_dimension[4].width = 4/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(4+1): column_dimension[4]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[5].auto_size = True
	column_dimension[5].width = 4/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(5+1): column_dimension[5]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[6].auto_size = True
	column_dimension[6].width = 4/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(6+1): column_dimension[6]})
	column_dimension.append(openpyxl.worksheet.ColumnDimension())
	column_dimension[7].auto_size = True
	column_dimension[7].width = 8/0.254		#cm
	f.column_dimensions.update({openpyxl.cell.get_column_letter(7+1): column_dimension[7]})
	row_dimension = []
	row_dimension.append(openpyxl.worksheet.RowDimension())
	row_dimension[0].height = 1.5*28	#cm
	f.row_dimensions.update({0+1: row_dimension[0]})
	row_dimension.append(openpyxl.worksheet.RowDimension())
	row_dimension[1].height = 1.0*28	#cm
	f.row_dimensions.update({1+1: row_dimension[1]})
	border = openpyxl.style.Borders()
	border.bottom.border_style = openpyxl.style.Border.BORDER_DOUBLE
	titre1 = openpyxl.style.Font()
	titre1.bold = True
	titre1.size = 18
	titre2 = openpyxl.style.Font()
	titre2.bold = True
	titre2.size = 14
	centre_centre = openpyxl.style.Alignment()
	centre_centre.horizontal = centre_centre.HORIZONTAL_CENTER
	centre_centre.vertical = centre_centre.VERTICAL_CENTER
	left = openpyxl.style.Alignment()
	left.horizontal = centre_centre.HORIZONTAL_LEFT
	# titre de la page
	f.cell(row=0, column=0).value = titre
	f.cell(row=0, column=0).style.alignment = centre_centre
	# entete de colonne
	f.cell(row=0, column=0).style.font = titre1
#	f.cell(row=1, column=0).value = "Dest."
	f.cell(row=1, column=0).style.borders = border
	f.cell(row=1, column=0).style.font = titre2
	f.cell(row=1, column=1).value = "Code barre"
	f.cell(row=1, column=1).style.borders = border
	f.cell(row=1, column=1).style.font = titre2
	f.cell(row=1, column=2).value = "Code"
	f.cell(row=1, column=2).style.borders = border
	f.cell(row=1, column=2).style.font = titre2
	f.cell(row=1, column=3).value = "Article"
	f.cell(row=1, column=3).style.borders = border
	f.cell(row=1, column=3).style.font = titre2
	f.cell(row=1, column=4).value = "Création"
	f.cell(row=1, column=4).style.borders = border
	f.cell(row=1, column=4).style.font = titre2
	f.cell(row=1, column=5).value = "Réception"
	f.cell(row=1, column=5).style.borders = border
	f.cell(row=1, column=5).style.font = titre2
	f.cell(row=1, column=6).value = "Suppression"
	f.cell(row=1, column=6).style.borders = border
	f.cell(row=1, column=6).style.font = titre2
	f.cell(row=1, column=7).value = "Emplacement"
	f.cell(row=1, column=7).style.borders = border
	f.cell(row=1, column=7).style.font = titre2
	row = 2
	for cont in liste:
		if type(cont) == types.DictType: # cont n'est pas un contenant mais un article
			f.cell(row=row, column=2).value = cont["nomenclature_code"]
			f.cell(row=row, column=2).style.alignment = left
			f.cell(row=row, column=3).value = cont["nomenclature_description"]
			row += 1
		elif type(cont) == Nomenclature: # cont n'est pas un contenant mais un article
			lot = cont.get_nomenclature_lot(supp=False)
			for l in lot:
				cc = l.get_nomenclature_lot_contenant(supp=False)
				for c in cc:
					f.cell(row=row, column=1).value = c.code
					f.cell(row=row, column=1).style.alignment = left
					f.cell(row=row, column=2).value = c.nomenclature_lot.nomenclature.code
					f.cell(row=row, column=2).style.alignment = left
					f.cell(row=row, column=3).value = c.nomenclature_lot.nomenclature.description
					f.cell(row=row, column=4).value = c.f_date_creation()
					f.cell(row=row, column=5).value = c.f_date_reception()
					f.cell(row=row, column=6).value = c.f_date_suppression()
					f.cell(row=row, column=7).value = c.actuel_emplacement_nom_court()
					row += 1
		else:
			f.cell(row=row, column=1).value = cont.code
			f.cell(row=row, column=1).style.alignment = left
			f.cell(row=row, column=2).value = cont.nomenclature_lot.nomenclature.code
			f.cell(row=row, column=2).style.alignment = left
			f.cell(row=row, column=3).value = cont.nomenclature_lot.nomenclature.description
			f.cell(row=row, column=4).value = cont.f_date_creation()
			f.cell(row=row, column=5).value = cont.f_date_reception()
			f.cell(row=row, column=6).value = cont.f_date_suppression()
			f.cell(row=row, column=7).value = cont.actuel_emplacement_nom_court()
			row += 1

def construction_xlsx(liste_recherche):
	marge_perdu = datetime.timedelta(days = MARGE_PERDU)
	marge_destruction = datetime.timedelta(days = MARGE_DESTRUCTION)
	marge_suppression = datetime.timedelta(days = MARGE_SUPPRESSION)
	marge_nouvelle_collection = datetime.timedelta(days = MARGE_NOUVELLE_COLLECTION)
	today = datetime.date.today()
	now = datetime.datetime.now()

	# Creation d'un fichier .xls vide
	xlsx = openpyxl.Workbook(encoding="utf-8")
	xlsx.properties.title = "Recherche pour néttoyage"
	xlsx.properties.category = "Auto export"
	xlsx.properties.company = "Linautom => Takasago"
	#xlsx.properties.created = datetime.datetime.now()
	xlsx.properties.creator = "Charly GONTERO"
	#xlsx.properties.description =
	#xlsx.properties.excel_base_date = datetime.datetime.now()
	#xlsx.properties.keywords =
	#xlsx.properties.last_modified_by = datetime.datetime.now()
	#xlsx.properties.modified =
	#xlsx.properties.subject =

	# Ajout des feuilles au fichier .xlsx
	xlsx.get_active_sheet()._set_title("Perdu") # renomme la feuille active
	xlsx.create_sheet(index=None, title="Validité inconnu") # crée de nouvelle feuille
	xlsx.create_sheet(index=None, title="Suppression destruction") # crée de nouvelle feuille
	xlsx.create_sheet(index=None, title="Destruction") # crée de nouvelle feuille
	xlsx.create_sheet(index=None, title="Fin de validité") # crée de nouvelle feuille
	xlsx.create_sheet(index=None, title="Collections épuisées") # crée de nouvelle feuille
	xlsx.create_sheet(index=None, title="Collections ajout") # crée de nouvelle feuille
	xlsx.create_sheet(index=None, title="Collections suppression") # crée de nouvelle feuille

	# Création d'un dico des feuille du fichier .xlsx
	log_print(xlsx.get_sheet_names())
	xlsx_sheet = {}
	for s in xlsx.get_sheet_names():
		xlsx_sheet.update({s : xlsx.get_sheet_by_name(s)})

	# Création pages dans fichier xlsx
	construction_page_xlsx(xlsx_sheet["Perdu"], liste_recherche["liste_perdu"], "Liste des contenants perdu ( le contenant n'est pas réceptionné depuis %d jours ) => destruction" %(marge_perdu.days))
	construction_page_xlsx(xlsx_sheet["Validité inconnu"], liste_recherche["liste_non_validite"], "Liste des contenants sans date de fin de validité mais réceptionné => destruction")
	construction_page_xlsx(xlsx_sheet["Suppression destruction"], liste_recherche["liste_suppression_destruction"], "Liste des contenants en fin de validité mais non supprimés ( validité + %d jours ) => destruction" %(marge_destruction.days))
	construction_page_xlsx(xlsx_sheet["Destruction"], liste_recherche["liste_destruction"], "Liste des contenants en fin de validité, supprimés ( validité + %d jours ) => destruction" %(marge_destruction.days))
	construction_page_xlsx(xlsx_sheet["Fin de validité"], liste_recherche["liste_fin_validite"], "Liste des contenants prêt pour destruction ( validité %d jours )" %(marge_suppression.days))
	construction_page_xlsx(xlsx_sheet["Collections épuisées"], liste_recherche["liste_collection_epuise"], "Liste des collections ne contenant pas de contenant valide")
	construction_page_xlsx(xlsx_sheet["Collections ajout"], liste_recherche["liste_collection_ajout"], "Liste des nouveaux article collections (Modif statut %d jours)" %(marge_nouvelle_collection.days))
	construction_page_xlsx(xlsx_sheet["Collections suppression"], liste_recherche["liste_collection_supp"], "Liste des articles plus en collection (Modif statut %d jours)" %(marge_nouvelle_collection.days))

	# Enregistrement du fichier exel pour envoie en pièce jointe
	fichier_xlsx = "%s/validite_%s.xlsx" %(XLSX_PATH, now.strftime("%Y-%m-%d_%H%M%S"))
	log_print("Sauvegarde du fichier : %s" %(fichier_xlsx))
	xlsx.save(fichier_xlsx)
	return {"xlsx": xlsx, "fichier_xlsx": fichier_xlsx}

def destruction(liste):
	user = User.objects.get(username="script")
	for cont in liste:
		log_print(cont)
		a = cont.destruction_test(user)
		if a != None:
			a.destruction(user)

def destruction_nomenclature_id(article_id):
	user = User.objects.get(username="script")
	article = Nomenclature.objects.get(pk=article_id)
	log_print(article_id, article)
	article.destruction(user)

def destruction_nomenclaturelot_id(lot_id):
	user = User.objects.get(username="script")
	lot = NomenclatureLot.objects.get(pk=lot_id)
	log_print(lot_id, lot)
	lot.destruction(user)

def recherche():
	log_print("Extraction ...")
	cont = Contenant.objects.all()
	marge_perdu = datetime.timedelta(days = MARGE_PERDU)
	marge_destruction = datetime.timedelta(days = MARGE_DESTRUCTION)
	marge_suppression = datetime.timedelta(days = MARGE_SUPPRESSION)
	marge_nouvelle_collection = datetime.timedelta(days = MARGE_NOUVELLE_COLLECTION)
	today = datetime.date.today()
	now = datetime.datetime.now()
	liste_perdu = []
	liste_non_validite = []
	liste_suppression_destruction = []
	liste_destruction = []
	liste_fin_validite = []
	liste_collection_ajout = []
	liste_collection_supp = []
	liste_lot_vide = []
	liste_article_vide = []
	for c in cont:
		if c.date_reception == None:
			if today - c.date_creation > marge_perdu:
				liste_perdu.append(c)
		elif c.date_fin_validite == None:
			liste_non_validite.append(c)
		elif c.date_fin_validite + marge_destruction < today:
			if c.date_suppression == None:
				liste_suppression_destruction.append(c)
			else:
				liste_destruction.append(c)
		elif c.date_fin_validite + marge_suppression < today:
			if c.date_suppression == None:
				liste_fin_validite.append(c)

	my=mysql.Mysql()
	my.open(mysql_host=settings.DATABASES["default"]["HOST"], mysql_port=int(settings.DATABASES["default"]["PORT"]), mysql_user=settings.DATABASES["default"]["USER"], mysql_password=settings.DATABASES["default"]["PASSWORD"] )

	# Liste des collections
	querry = """
		SELECT stock_labo_nomenclature.code as nomenclature_code, stock_labo_nomenclature.description as nomenclature_description
		FROM %s.stock_labo_nomenclature
		WHERE stock_labo_nomenclature.collection = 1""" %(settings.DATABASES["default"]["NAME"])
	liste_collection = [i for i in my.execute(querry)]

	# Liste des collections aillant un contenant valide
	querry = """
		SELECT stock_labo_contenant.code as contenant_code,
			stock_labo_nomenclature.code as nomenclature_code,
			stock_labo_nomenclature.description as nomenclature_description,
			stock_labo_contenant.date_creation as contenant_date_creation,
			stock_labo_contenant.date_reception as contenant_date_reception,
			stock_labo_contenant.date_suppression as contenant_date_suppression,
			stock_labo_contenant.date_fin_validite as contenant_date_fin_validite,

			stock_labo_stocksite.description as stocksite_description,
			stock_labo_stockentrepot.description as labo_stockentrepot_description,
			stock_labo_stockmagasin.description as stockmagasin_description

		FROM %s.stock_labo_contenant 
			LEFT JOIN %s.stock_labo_nomenclaturelot ON stock_labo_contenant.nomenclature_lot_id = stock_labo_nomenclaturelot.id
			LEFT JOIN %s.stock_labo_nomenclature ON stock_labo_nomenclaturelot.nomenclature_id = stock_labo_nomenclature.id

			LEFT JOIN %s.stock_labo_stocksite ON stock_labo_contenant.actuel_site_id = stock_labo_stocksite.id
			LEFT JOIN %s.stock_labo_stockentrepot ON stock_labo_contenant.actuel_entrepot_id = stock_labo_stockentrepot.id
			LEFT JOIN %s.stock_labo_stockmagasin ON stock_labo_contenant.actuel_magasin_id = stock_labo_stockmagasin.id

		WHERE stock_labo_contenant.date_suppression IS NULL and stock_labo_nomenclature.collection = 1
		GROUP BY stock_labo_nomenclature.code"""  %(settings.DATABASES["default"]["NAME"],
													settings.DATABASES["default"]["NAME"],
													settings.DATABASES["default"]["NAME"],
													settings.DATABASES["default"]["NAME"],
													settings.DATABASES["default"]["NAME"],
													settings.DATABASES["default"]["NAME"])
	contenant = my.execute(querry)
	# Liste des nomenclature_lot ne contenant plus de contenant
	querry = """
		SELECT	`stock_labo_nomenclaturelot`.`id` as `nomenclaturelot_id`,
				`stock_labo_nomenclaturelot`.`code` as `nomenclaturelot_code`
		FROM `%s`.`stock_labo_nomenclaturelot` 
		LEFT JOIN `%s`.`stock_labo_contenant` ON `stock_labo_contenant`.`nomenclature_lot_id` = `stock_labo_nomenclaturelot`.`id`
		WHERE `stock_labo_contenant`.`code` IS NULL
		"""  %(settings.DATABASES["default"]["NAME"],
			settings.DATABASES["default"]["NAME"])
	liste_lot_vide = my.execute(querry)
	# Liste des nomenclature ne contennant plus de lot
		#if self.collection == True:
		#	a = None
		#elif self.duree_validite == settings_default.DUREE_VALIDITE_MP:
		#	if self.get_nomenclature_lot().count() == 1:
		#		a = self
		#	else:
		#		a = None
		#elif self.duree_validite == settings_default.DUREE_VALIDITE_COUP:
		#	if self.get_nomenclature_lot().count() == 1:
		#		a = self
		#	else:
		#		a = None
		#else:
		#	a = None
	querry = """
		SELECT `stock_labo_nomenclature`.`id` as `nomenclature_id`,  
			`stock_labo_nomenclature`.`code` as `nomenclature_code`,
			`stock_labo_nomenclature`.`description` as `nomenclature_description`,
			`stock_labo_nomenclaturelot`.`code` as `nomenclaturelot_code`
		FROM `%s`.`stock_labo_nomenclature` 
		LEFT JOIN `%s`.`stock_labo_nomenclaturelot` ON `stock_labo_nomenclaturelot`.`nomenclature_id` = `stock_labo_nomenclature`.`id`
		WHERE `stock_labo_nomenclaturelot`.`code` IS NULL
			AND `stock_labo_nomenclature`.`collection` = "0"
			AND ( `stock_labo_nomenclature`.`duree_validite` = "%s" or `stock_labo_nomenclature`.`duree_validite` = "%s" )
		"""  %(settings.DATABASES["default"]["NAME"],
			settings.DATABASES["default"]["NAME"],
			DUREE_VALIDITE_MP,
			DUREE_VALIDITE_COUP)
	liste_article_vide = my.execute(querry)
	my.close()

	for cont in contenant:
		i = 0
		while i < len(liste_collection):
			if liste_collection[i]["nomenclature_code"] == cont["nomenclature_code"]:
				liste_collection.pop(i)
				break
			i += 1

	# Liste des entrée et sortie de collection à moins de 30 jours
	list_n = Nomenclature.objects.filter(date_collection__gt=today-marge_nouvelle_collection)
	for n in list_n:
		if n.collection == True:
			liste_collection_ajout.append(n)
		else:
			liste_collection_supp.append(n)
			
	log_print("Nombre de contenant total : %d" %(Contenant.objects.all().count()))
	log_print("Liste des perdu : %s" %(len(liste_perdu)))
	log_print("Liste validité inconnu : %s" %(len(liste_non_validite)))
	log_print("Liste des suppression destruction : %s" %(len(liste_suppression_destruction)))
	log_print("Liste des destruction : %s" %(len(liste_destruction)))
	log_print("Liste des Fin de validité : %s" %(len(liste_fin_validite)))
	log_print("Liste collection épuisée : %s" %(len(liste_collection)))
	log_print("Liste collection ajoutée : %s" %(len(liste_collection_ajout)))
	log_print("Liste collection supprimé : %s" %(len(liste_collection_supp)))
	log_print("Liste lot_vide : %s" %(len(liste_lot_vide)))
	log_print("Liste article_vide : %s" %(len(liste_article_vide)))


	
	return {
		"liste_perdu":liste_perdu,
		"liste_non_validite":liste_non_validite,
		"liste_suppression_destruction":liste_suppression_destruction,
		"liste_destruction":liste_destruction,
		"liste_fin_validite":liste_fin_validite,
		"liste_collection_epuise":liste_collection,
		"liste_collection_ajout":liste_collection_ajout,
		"liste_collection_supp":liste_collection_supp,
		"liste_lot_vide":liste_lot_vide,
		"liste_article_vide":liste_article_vide,
		}




def script_nettoyage():
	"""
	Script de nettoyage , recherche des contenant
		Liste des perdu
		Liste des validité inconnu
		Liste des suppression destruction
		Liste des destruction
		Liste des Fin de validité
	Envoie email fichier exel
	Execution destruction des contenants
		Liste des perdu
		Liste des validité inconnu
		Liste des suppression destruction
		Liste des destruction
	"""
	liste_recherche = recherche()

	dico_fichier_xlsx = construction_xlsx(liste_recherche)

	# Envoie email
	mail_corps = "Les contenants dans les feuilles Perdu, Validité inconnu, Supperssion destruction et Destruction sont détruit pour nettoyage\r\n\r\nLes contenants de la feuille Fin de validité sont les contenants arrivant en fin de validité"
	m = mail.MailAttach(mail_corps)
	m.from_("Application_stockparis@takasago.com")
	m.reply_to("charly.gontero@linautom.fr")
	m.to(["charly.gontero@linautom.fr","yanick_soufflet@takasago.com","TEPL-gestion-pt@takasago.com"])
	if version.status_developement["data"] == "prod":
		m.subject("Stock Labo Paris, Nettoyage et fin de validité")
	else:
		m.subject("Ne pas tenir compte de cette email, je fait des tests sur le serveur, charly")
	m.attach(dico_fichier_xlsx["fichier_xlsx"])
	log_print("Mail to %s ... " %(m.list_to))
#	m.send(smtp="smtp.laposte.net", expediteur="charly.gontero@laposte.net", user="charly.gontero", password="1Addyson", starttls=False)
	m.send(smtp="par-srv-cas01.eu.takasago.com", expediteur="Application_stockparis@takasago.com", user=None , password=None, starttls=False)

	# Execution destruction
	log_print("Destruction de la liste : liste_perdu")
	destruction(liste_recherche["liste_perdu"])
	log_print("Destruction de la liste : liste_non_validite")
	destruction(liste_recherche["liste_non_validite"])
	log_print("Destruction de la liste : liste_suppression_destruction")
	destruction(liste_recherche["liste_suppression_destruction"])
	log_print("Destruction de la liste : liste_destruction")
	destruction(liste_recherche["liste_destruction"])

	log_print("Liste des articles ne contenant plus de lot et qui doivent être supprimé")
	for a in liste_recherche["liste_article_vide"]:
		destruction_nomenclature_id(a["nomenclature_id"])
	log_print ("Liste des lots qui contenant plus de contenant et dui doivent être supprimmé")
	for l in liste_recherche["liste_lot_vide"]:
		destruction_nomenclaturelot_id(l["nomenclaturelot_id"])

if __name__ == '__main__':

	script_nettoyage()


